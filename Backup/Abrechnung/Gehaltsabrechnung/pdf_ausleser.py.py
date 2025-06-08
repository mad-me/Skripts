import os
import sys
import subprocess
import re
import pdfplumber
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers

# Pakete sicherstellen
def install(package):
    subprocess.check_call([sys.executable, "-m", "pip", "install", package])

try:
    import pdfplumber
    import pandas as pd
    import openpyxl
except ImportError:
    install("pdfplumber")
    install("pandas")
    install("openpyxl")
    import pdfplumber
    import pandas as pd
    import openpyxl

# Ordner
downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")
output_folder = "Excel"
if not os.path.exists(output_folder):
    os.makedirs(output_folder)

# Neueste passende Datei im Downloads-Ordner finden
pdf_dateien = [f for f in os.listdir(downloads_path) if re.match(r"Abrechnungen \d{2}_\d{4}\.pdf", f)]
if not pdf_dateien:
    print("❌ Keine passende PDF-Datei im Downloads-Ordner gefunden!")
    sys.exit(1)

pdf_dateien.sort(key=lambda f: os.path.getmtime(os.path.join(downloads_path, f)), reverse=True)
neuste_pdf = pdf_dateien[0]
pdf_path = os.path.join(downloads_path, neuste_pdf)

print(f"🔍 Verwende Datei: {neuste_pdf}")

alle_daten = []
monat_gefunden = None
erfolgreich_verarbeitet = False

with pdfplumber.open(pdf_path) as pdf:
    for page in pdf.pages:
        text = page.extract_text()
        if not text:
            continue

        if not monat_gefunden:
            monat_match = re.search(r"Monat:\s*(\d{2}/\d{4})", text)
            if monat_match:
                monat_gefunden = monat_match.group(1).replace('/', '_')

        name = re.search(r"Dienstnehmer:\s+(.*?)\s+DN-Nr", text)
        brutto = re.search(r"Brutto\s+([\d,.]+)\s+EURO", text)
        sv = re.search(r"SV-Beiträge:\s+([\d,.]+)", text)
        netto = re.search(r"Zahlbetrag\s+([\d,.]+)\s+EURO", text)

        if name and brutto and sv and netto:
            brutto_eur = float(brutto.group(1).replace('.', '').replace(',', '.'))
            sv_eur = float(sv.group(1).replace('.', '').replace(',', '.'))
            netto_eur = float(netto.group(1).replace('.', '').replace(',', '.'))

            alle_daten.append({
                "Name": name.group(1).strip(),
                "Brutto (€)": brutto_eur,
                "SV-Abgaben (€)": sv_eur,
                "Netto (€)": netto_eur,
                "48,96% von Brutto (€)": round(brutto_eur * 0.4896, 2)
            })

            erfolgreich_verarbeitet = True

# Dynamischer Dateiname
if monat_gefunden:
    dateiname = f"Abrechnungen_{monat_gefunden}.xlsx"
else:
    dateiname = "Abrechnungen_gesamt.xlsx"

# Speichern
output_path = os.path.join(output_folder, dateiname)
df = pd.DataFrame(alle_daten)
df.to_excel(output_path, index=False)

# Jetzt Formatierungen mit openpyxl
wb = load_workbook(output_path)
ws = wb.active

# Automatische Spaltenbreite
for column in ws.columns:
    max_length = 0
    column_letter = get_column_letter(column[0].column)
    for cell in column:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column_letter].width = adjusted_width

# Spalten B bis E als Buchhaltungsformat formatieren
for col_letter in ['B', 'C', 'D', 'E']:
    for cell in ws[col_letter][1:]:  # ab Zeile 2 (keine Überschrift)
        cell.number_format = '"€"* #,##0.00_);[Red]("€"* #,##0.00)'

wb.save(output_path)
wb.close()

# Optionale Löschung
if erfolgreich_verarbeitet:
    os.remove(pdf_path)
    print(f"🗑️ Datei {neuste_pdf} wurde nach Verarbeitung gelöscht.")
else:
    print(f"⚠️ Datei {neuste_pdf} konnte nicht vollständig verarbeitet werden und wurde NICHT gelöscht!")

print(f"✅ Fertig! Datei gespeichert unter: {output_path}")
input("Drücke Enter zum Beenden...")
