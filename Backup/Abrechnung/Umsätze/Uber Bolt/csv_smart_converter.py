import pandas as pd
import os
import re
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from pathlib import Path

# ➡️ Einstellungen
output_folder = "Excel"
downloads_folder = str(Path.home() / "Downloads")

os.makedirs(output_folder, exist_ok=True)

all_dataframes = []
erste_kw = None
output_kw_folder = output_folder

# ➡️ Funktion: Suche die jeweils neueste Bolt- und Uber-Datei
def find_specific_files(downloads_folder):
    bolt_files = []
    uber_files = []
    for filename in os.listdir(downloads_folder):
        if filename.lower().endswith(".csv"):
            if "umsatzbericht" in filename.lower() and "el kaptin kg" in filename.lower():
                bolt_files.append(os.path.join(downloads_folder, filename))
            if "driver_performance" in filename.lower() and "el_kaptin_kg" in filename.lower():
                uber_files.append(os.path.join(downloads_folder, filename))
    bolt_file = max(bolt_files, key=os.path.getmtime) if bolt_files else None
    uber_file = max(uber_files, key=os.path.getmtime) if uber_files else None
    return bolt_file, uber_file

# ➡️ Funktion: Kalenderwoche aus Dateiname extrahieren
def extract_kw_from_filename(filename):
    match_direct_kw = re.search(r'(\d{4})W(\d{2})', filename)
    if match_direct_kw:
        kw = int(match_direct_kw.group(2))
        return f"KW{kw}"

    match_uber = re.search(r'(\d{8})-(\d{8})', filename)
    if match_uber:
        start_date = datetime.strptime(match_uber.group(1), "%Y%m%d")
        kw = start_date.isocalendar()[1]
        return f"KW{kw}"

    match_bolt = re.search(r'(\d{2})_(\d{2})_(\d{4})-(\d{2})_(\d{2})_(\d{4})', filename)
    if match_bolt:
        day, month, year = int(match_bolt.group(1)), int(match_bolt.group(2)), int(match_bolt.group(3))
        start_date = datetime(year, month, day)
        kw = start_date.isocalendar()[1]
        return f"KW{kw}"

    return "Unbekannt"

# ➡️ Funktion: Saubere eindeutige Dateinamen erzeugen
def generate_unique_filename(base_path, base_name):
    counter = 1
    file_path = os.path.join(base_path, base_name)
    while os.path.exists(file_path):
        name_without_ext, ext = os.path.splitext(base_name)
        file_path = os.path.join(base_path, f"{name_without_ext}_{counter}{ext}")
        counter += 1
    return file_path

# ➡️ Bolt und Uber Verarbeitung
def process_bolt(df):
    df['Name'] = df['Fahrer'].str.strip()
    df['Netto Umsatz (€)'] = df['Netto-Einnahmen|€']
    df['Bargeld erhalten (€)'] = df['Bargeld erhalten|€']
    df['Differenz (€)'] = df['Netto Umsatz (€)'] - df['Bargeld erhalten (€)']
    df['Plattform'] = 'Bolt'
    return df[['Name', 'Netto Umsatz (€)', 'Bargeld erhalten (€)', 'Differenz (€)', 'Plattform']]

def process_uber(df):
    df['Name'] = df['Vorname des Fahrers'].str.strip() + ' ' + df['Nachname des Fahrers'].str.strip()
    df['Netto Umsatz (€)'] = df['Gesamtumsätze']
    df['Bargeld erhalten (€)'] = df['Eingenommenes Bargeld']
    df['Differenz (€)'] = df['Netto Umsatz (€)'] - df['Bargeld erhalten (€)']
    df['Plattform'] = 'Uber'
    return df[['Name', 'Netto Umsatz (€)', 'Bargeld erhalten (€)', 'Differenz (€)', 'Plattform']]

# ➡️ Starte: Suche die neuesten passenden Dateien
bolt_file, uber_file = find_specific_files(downloads_folder)

# ➡️ Verarbeite beide Dateien
for csv_path in [bolt_file, uber_file]:
    if csv_path:
        filename = os.path.basename(csv_path)

        try:
            df = pd.read_csv(csv_path, encoding='utf-8-sig', delimiter=',')
        except Exception as e:
            print(f"❌ Fehler beim Lesen von {filename}: {e}")
            continue

        print(f"🔎 Verarbeite Datei: {filename}")

        if erste_kw is None:
            erste_kw = extract_kw_from_filename(filename)
            if erste_kw and erste_kw != "Unbekannt":
                output_kw_folder = os.path.join(output_folder, erste_kw)
                os.makedirs(output_kw_folder, exist_ok=True)

        if 'Fahrer' in df.columns and 'Netto-Einnahmen|€' in df.columns:
            print("➡️ Plattform erkannt: BOLT")
            df_result = process_bolt(df)
            base_filename = f"Bolt_{erste_kw}.xlsx" if erste_kw else "Bolt_Unbekannt.xlsx"
        elif 'Vorname des Fahrers' in df.columns and 'Gesamtumsätze' in df.columns and 'Eingenommenes Bargeld' in df.columns:
            print("➡️ Plattform erkannt: UBER")
            df_result = process_uber(df)
            base_filename = f"Uber_{erste_kw}.xlsx" if erste_kw else "Uber_Unbekannt.xlsx"
        else:
            print(f"⚠️ Datei {filename} konnte nicht automatisch erkannt werden.")
            continue

        excel_path = generate_unique_filename(output_kw_folder, base_filename)
        df_result.to_excel(excel_path, index=False)

        all_dataframes.append(df_result)

        try:
            os.remove(csv_path)
            print(f"🗑️ Datei {filename} erfolgreich aus Downloads gelöscht.")
        except Exception as e:
            print(f"⚠️ Fehler beim Löschen der Datei {filename}: {e}")

# ➡️ Sammeldatei erstellen
if all_dataframes:
    print("📊 Erstelle Sammeldatei...")
    gesamtdaten = pd.concat(all_dataframes)
    gesamtdaten = gesamtdaten.sort_values(by=["Name"])
    gesamtdaten.reset_index(drop=True, inplace=True)

    gesamtdaten = gesamtdaten[(gesamtdaten['Netto Umsatz (€)'] != 0) | (gesamtdaten['Bargeld erhalten (€)'] != 0)]

    gesamtdaten_filename = f"Gesamtübersicht_{erste_kw}.xlsx" if erste_kw else "Gesamtübersicht_Unbekannt.xlsx"
    gesamtdaten_path = os.path.join(output_kw_folder, gesamtdaten_filename)

    gesamtdaten.to_excel(gesamtdaten_path, index=False)

    # ➡️ Formatierungen
    wb = load_workbook(gesamtdaten_path)
    ws = wb.active

    # Spaltenbreite
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2)
        ws.column_dimensions[col_letter].width = adjusted_width

    # Zoom auf 265%
    ws.sheet_view.zoomScale = 265

    # Alternierende Zeilenfarben
    fill_grau = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
    for idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        if idx % 2 == 0:
            for cell in row:
                cell.fill = fill_grau

    # Währungsformat für B, C, D
    for col_letter in ['B', 'C', 'D']:
        for cell in ws[col_letter][1:]:
            cell.number_format = '#,##0.00 €'

    wb.save(gesamtdaten_path)
    print(f"✅ Sammeldatei perfekt gespeichert: {gesamtdaten_path}")

print("✅ Alle CSV-Dateien wurden erfolgreich verarbeitet!")
