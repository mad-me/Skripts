import os
import glob
import re
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter

# === KONFIGURATION ===
DOWNLOAD_FOLDER = os.path.expanduser('~/Downloads')  # Pfad zum Download-Ordner anpassen

# 1. Neueste Datei suchen (.belegliste_YYYY.MM.DD_hhmm_YYYY.MM.DD_hhmm.csv)
pattern = os.path.join(DOWNLOAD_FOLDER, '*.belegliste_*.csv')
files = glob.glob(pattern)
if not files:
    raise FileNotFoundError(f"Keine Dateien gefunden mit Muster: {pattern}")
# Filter nach Datums-Pattern im Dateinamen
date_files = []
for f in files:
    name = os.path.basename(f).lower()
    m = re.search(r'belegliste_(\d{4})\.(\d{2})\.(\d{2})_\d{4}_(\d{4})\.(\d{2})\.(\d{2})_\d{4}\.csv$', name)
    if m:
        date_files.append((f, m))
if not date_files:
    raise ValueError("Keine belegliste-Dateien im erwarteten Datumsformat gefunden.")
# Neueste Datei wählen
date_files.sort(key=lambda x: os.path.getmtime(x[0]), reverse=True)
latest_file, latest_match = date_files[0]
print(f"Verwendete Datei: {latest_file}")

# 2. Kalenderwoche anhand Start-Datum
year1, month1, day1 = map(int, latest_match.groups()[0:3])
date_start = datetime(year1, month1, day1)
kw = date_start.isocalendar()[1]

# 3. Ausgabe-Datei
output_folder = r"D:\Abrechnung\Umsätze\40100\Excel"
os.makedirs(output_folder, exist_ok=True)
output_file = os.path.join(output_folder, f"40100_KW{kw:02d}.xlsx")
print(f"Ausgabe: {output_file}")

# 4. CSV einlesen
df = pd.read_csv(latest_file, sep=';', encoding='utf-8')

# 5. Spalten G–L numerisch konvertieren (Index 6–11)
cols_GL = df.columns[6:12]
df[cols_GL] = df[cols_GL].apply(lambda col: pd.to_numeric(
    col.astype(str)
       .str.replace('.', '', regex=False)
       .str.replace(',', '.', regex=False),
    errors='coerce'
))

# 6. Excel-Ausgabe pro Fahrzeug
driver = pd.ExcelWriter(output_file, engine='openpyxl')
with driver as writer:
    sheet_names = set()
    for vehicle, group in df.groupby('Fahrzeug'):
        # Blattname bereinigen und eindeutig machen
        name = re.sub(r"[\[\]\*:/\\?]", "", str(vehicle))[:31]
        base, count = name, 1
        while name in sheet_names:
            suffix = f"_{count}"
            name = f"{base[:31-len(suffix)]}{suffix}"
            count += 1
        sheet_names.add(name)
        group.to_excel(
            writer,
            sheet_name=name,
            index=False,
            header=False,
            startrow=3
        )

# 7. Formatierung in Openpyxl
wb = openpyxl.load_workbook(output_file)
for ws in wb.worksheets:
    # 7.1. Header in Zeile 1
    for idx, col in enumerate(df.columns, start=1):
        ws.cell(row=1, column=idx, value=col)
    max_row = ws.max_row

    # 7.2. SUMIF-Formeln
    # E3: Summen aller G ab Zeile 4, D != 'Barbeleg'
    formula_E3 = (
        f'=SUMIF({get_column_letter(4)}4:'
        f'{get_column_letter(4)}{max_row},"<>Barbeleg",'
        f'{get_column_letter(7)}4:{get_column_letter(7)}{max_row})'
    )
    ws.cell(row=3, column=5).value = formula_E3
    # D3: Summen aller G ab Zeile 4, D = 'Barbeleg'
    formula_D3 = (
        f'=SUMIF({get_column_letter(4)}4:'
        f'{get_column_letter(4)}{max_row},"=Barbeleg",'
        f'{get_column_letter(7)}4:{get_column_letter(7)}{max_row})'
    )
    ws.cell(row=3, column=4).value = formula_D3

        # 7.3. Summenzeile 3: SUM in G–L (7–12) in Zeile 3
    for col_idx in range(7, 13):
        letter = get_column_letter(col_idx)
        # Summenformel in derselben Zeile wie D3/E3
        ws.cell(row=3, column=col_idx).value = f'=SUM({letter}4:{letter}{max_row})'

    # 7.4. Spalten M–P löschen (Index 13–16). Spalten M–P löschen (Index 13–16)
    for col_idx in range(16, 12, -1):
        ws.delete_cols(col_idx)

    # 7.5. Buchhaltungsformat für D3 und E3
    for coord in [('D', 3), ('E', 3)]:
        ws[f"{coord[0]}{coord[1]}"] = ws[f"{coord[0]}{coord[1]}"].value
        ws[f"{coord[0]}{coord[1]}"] .number_format = '_-* #,##0.00 €_-'

    # 7.6. Spalten G–L als Buchhaltung ab Zeile 3
    for row in ws.iter_rows(min_row=3, max_row=max_row, min_col=7, max_col=12):
        for cell in row:
            cell.number_format = '_-* #,##0.00 €_-'

    # 7.7. Optimale Spaltenbreite
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                length = len(str(cell.value))
                if length > max_length:
                    max_length = length
        ws.column_dimensions[col_letter].width = max_length + 2

# 8. Speichern\ nwb.save(output_file)
wb.save(output_file)
print("Fertig!")
